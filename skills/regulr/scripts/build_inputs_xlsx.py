"""Build ``pricing_inputs.xlsx`` from a JSON spec.

Usage::

    python build_inputs_xlsx.py <out_dir> <spec.json>

Writes ``<out_dir>/pricing_inputs.xlsx`` with five sheets matching the schema
``pricing.py`` expects.

JSON spec shape::

    {
      "latest_data": [
        {"AccidentYear": 2016, "EarnedPremium": ..., "OnLevelFactor": ...,
         "PaidLoss": ..., "ReportedLoss": ..., "LossTrend": ...,
         "ExposureTrend": ...},
        ...
      ],
      "ldf_inputs": [
        {"AccidentYear": 2016, "UserPaid": ..., "UserReported": ...,
         "IndustryPaid": ..., "IndustryReported": ...,
         "BrokerPaid": ..., "BrokerReported": ...},
        ...
      ],
      "triangle_paid":     {"dev_ages": [12, 24, ...],
                            "rows": [{"AccidentYear": 2016,
                                      "values": [300000, 550000, ...]}, ...]},
      "triangle_reported": {"dev_ages": [...], "rows": [...]},
      "assumptions": {"TargetEffectiveDate": "2026-07-01",
                      "TailFactor_Paid": 1.005,
                      "TailFactor_Reported": 1.002,
                      "DevAges": "12,24,36,48,60,72,84,96,108,120"}
    }

``triangle_paid`` and ``triangle_reported`` are optional — omit (or set to
``null``) to skip those sheets.
"""
from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

LATEST_COLS = [
    "AccidentYear", "EarnedPremium", "OnLevelFactor",
    "PaidLoss", "ReportedLoss", "LossTrend", "ExposureTrend",
]
LDF_COLS = [
    "AccidentYear", "UserPaid", "UserReported",
    "IndustryPaid", "IndustryReported", "BrokerPaid", "BrokerReported",
]
ASSUMPTION_ORDER = [
    "TargetEffectiveDate", "TailFactor_Paid", "TailFactor_Reported", "DevAges",
]


def _write_records(ws, columns, records):
    ws.append(columns)
    for rec in records:
        ws.append([rec.get(c) for c in columns])


def _write_triangle(ws, spec):
    dev_ages = spec["dev_ages"]
    ws.append(["AccidentYear", *dev_ages])
    for row in spec["rows"]:
        values = row["values"]
        padded = list(values) + [None] * (len(dev_ages) - len(values))
        ws.append([row["AccidentYear"], *padded])


def _coerce_date(value):
    if isinstance(value, (date, datetime)):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return value
    return value


def _write_assumptions(ws, asmp):
    ws.append(["Parameter", "Value"])
    seen = set()
    for key in ASSUMPTION_ORDER:
        if key in asmp:
            value = asmp[key]
            if key == "TargetEffectiveDate":
                value = _coerce_date(value)
            ws.append([key, value])
            seen.add(key)
    for key, value in asmp.items():
        if key not in seen:
            ws.append([key, value])


def build(spec: dict, out_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    if spec.get("triangle_paid"):
        _write_triangle(wb.create_sheet("triangle_paid"), spec["triangle_paid"])
    if spec.get("triangle_reported"):
        _write_triangle(wb.create_sheet("triangle_reported"), spec["triangle_reported"])

    _write_records(wb.create_sheet("latest_data"), LATEST_COLS, spec["latest_data"])
    _write_records(wb.create_sheet("ldf_inputs"), LDF_COLS, spec["ldf_inputs"])
    _write_assumptions(wb.create_sheet("assumptions"), spec["assumptions"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main(argv):
    if len(argv) != 3:
        print("usage: build_inputs_xlsx.py <out_dir> <spec.json>", file=sys.stderr)
        sys.exit(2)
    out_dir = Path(argv[1]).resolve()
    spec_path = Path(argv[2]).resolve()
    spec = json.loads(spec_path.read_text(encoding="utf-8"))
    out_path = out_dir / "pricing_inputs.xlsx"
    build(spec, out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main(sys.argv)
