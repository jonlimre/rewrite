"""Actuarial pricing analysis — input loader and analysis-workbook generator.

Reads ``pricing_inputs.xlsx`` from the same folder as this script.  Sheets:

* ``triangle_paid``, ``triangle_reported`` — AY x dev age, both optional.
* ``latest_data`` — one row per AY (AccidentYear, EarnedPremium, OnLevelFactor,
  PaidLoss, ReportedLoss, LossTrend, ExposureTrend).  Each AY's current dev age
  is derived from its position (oldest → terminal age).
* ``ldf_inputs`` — one row per AY with a single age-to-age factor per source
  (UserPaid, UserReported, IndustryPaid, IndustryReported, BrokerPaid,
  BrokerReported).  Each factor is the AY's current-age-to-next-age step.  The
  Per-AY LDF table picks a Selected ATA per AY and cumulates to ultimate as the
  product down the Selected column.
* ``assumptions`` — Parameter/Value: TargetEffectiveDate, TailFactor_Paid,
  TailFactor_Reported, DevAges (comma-separated list, e.g. "12,24,36,...,120").

Running the script builds ``pricing_analysis.xlsx`` next to the input, populated
entirely with Excel formulas so the analysis is auditable and editable.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

WORKBOOK = "pricing_inputs.xlsx"
ANALYSIS_WORKBOOK = "pricing_analysis.xlsx"
REPORT_PDF = "pricing_report.pdf"
TRIANGLE_SHEETS = {"paid": "triangle_paid", "reported": "triangle_reported"}
LATEST_SHEET = "latest_data"
LDF_INPUTS_SHEET = "ldf_inputs"
ASSUMPTIONS_SHEET = "assumptions"


def _path(name: str = WORKBOOK) -> Path:
    return Path(__file__).resolve().parent / name


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------
def load_latest(path: Path | None = None) -> pd.DataFrame:
    return pd.read_excel(path or _path(), sheet_name=LATEST_SHEET)


def _melt_triangle(df: pd.DataFrame, kind: str) -> pd.DataFrame:
    df = df.rename(columns={df.columns[0]: "ay"}).melt(
        id_vars="ay", var_name="dev_age", value_name="value"
    )
    df["dev_age"] = pd.to_numeric(df["dev_age"], errors="coerce").astype("Int64")
    df["type"] = kind
    return df.dropna(subset=["value"]).reset_index(drop=True)


def load_triangles(path: Path | None = None) -> pd.DataFrame:
    book = pd.ExcelFile(path or _path())
    parts = []
    for kind, sheet in TRIANGLE_SHEETS.items():
        if sheet in book.sheet_names:
            raw = pd.read_excel(book, sheet_name=sheet)
            if not raw.empty:
                parts.append(_melt_triangle(raw, kind))
    if not parts:
        return pd.DataFrame(columns=["ay", "dev_age", "value", "type"])
    return pd.concat(parts, ignore_index=True)


def load_ldf_inputs(path: Path | None = None) -> pd.DataFrame:
    return pd.read_excel(path or _path(), sheet_name=LDF_INPUTS_SHEET)


def load_assumptions(path: Path | None = None) -> dict:
    df = pd.read_excel(path or _path(), sheet_name=ASSUMPTIONS_SHEET)
    asmp = dict(zip(df["Parameter"], df["Value"]))
    if isinstance(asmp.get("DevAges"), str):
        asmp["DevAges"] = [int(s.strip()) for s in asmp["DevAges"].split(",") if s.strip()]
    return asmp


def load_all(path: Path | None = None):
    p = path or _path()
    return load_latest(p), load_triangles(p), load_ldf_inputs(p), load_assumptions(p)


# ---------------------------------------------------------------------------
# Analysis-workbook generator
# ---------------------------------------------------------------------------
BOLD = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor="DDDDDD")


def _write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns):
        c = ws.cell(start_row, start_col + j, col)
        c.font = BOLD; c.fill = HDR_FILL
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, v in enumerate(row):
            if isinstance(v, pd.Timestamp):
                v = v.to_pydatetime()
            elif isinstance(v, list):
                v = ",".join(str(x) for x in v)
            elif not isinstance(v, str) and pd.isna(v):
                v = None
            ws.cell(i, start_col + j, v)


_ASMP_CANONICAL_ORDER = [
    "TargetEffectiveDate",
    "TailFactor_Paid",
    "TailFactor_Reported",
    "DevAges",
]


def _write_inputs(wb, latest, ldf_inputs, asmp, triangles):
    _write_df(wb.create_sheet("Inputs_Latest"), latest)
    _write_df(wb.create_sheet("Inputs_LDFs"), ldf_inputs)
    # Force canonical row order so hardcoded $B$2/$B$3/$B$4 formula refs
    # always resolve to TargetEffectiveDate / TailFactor_Paid / TailFactor_Reported
    # regardless of how the user ordered rows in the input file.
    ordered = [(k, asmp[k]) for k in _ASMP_CANONICAL_ORDER if k in asmp]
    extras = [(k, v) for k, v in asmp.items() if k not in _ASMP_CANONICAL_ORDER]
    asmp_df = pd.DataFrame(ordered + extras, columns=["Parameter", "Value"])
    _write_df(wb.create_sheet("Inputs_Assumptions"), asmp_df)
    for kind in TRIANGLE_SHEETS:
        sub = triangles[triangles["type"] == kind]
        if sub.empty:
            continue
        wide = (
            sub.pivot(index="ay", columns="dev_age", values="value")
            .sort_index().reset_index().rename(columns={"ay": "AccidentYear"})
        )
        _write_df(wb.create_sheet(f"Inputs_Triangle_{kind.capitalize()}"), wide)


# ---------------------------------------------------------------------------
# LDF sheet
# ---------------------------------------------------------------------------
_LDF_SRC = {
    "paid":     {"User": "B", "Industry": "D", "Broker": "F", "TailRow": 3},
    "reported": {"User": "C", "Industry": "E", "Broker": "G", "TailRow": 4},
}


def _build_ldf_sheet(wb, kind, latest, has_tri, dev_ages):
    ws = wb.create_sheet(f"LDFs_{kind.capitalize()}")
    tri_sheet = f"Inputs_Triangle_{kind.capitalize()}"
    n_ays = len(latest)
    src = _LDF_SRC[kind]
    pair_labels = [f"{a}-{b}" for a, b in zip(dev_ages[:-1], dev_ages[1:])]
    n_pairs = len(pair_labels)
    tail_col = 2 + n_pairs

    # Row 1: headers
    ws.cell(1, 1, "Development").font = BOLD
    for k, lab in enumerate(pair_labels):
        ws.cell(1, 2 + k, lab).font = BOLD
    ws.cell(1, tail_col, "Tail").font = BOLD

    ws.cell(2, 1, "Age-to-Age by AY:").font = BOLD

    if has_tri:
        for i in range(n_ays):
            r = 3 + i
            ws.cell(r, 1, f"=Inputs_Latest!A{2 + i}")
            for k in range(n_pairs):
                f_col = get_column_letter(2 + k)
                t_col = get_column_letter(3 + k)
                ws.cell(r, 2 + k,
                    f"=IF(ISBLANK({tri_sheet}!{t_col}{2 + i}),\"\","
                    f"{tri_sheet}!{t_col}{2 + i}/{tri_sheet}!{f_col}{2 + i})")

    # Averages
    ws.cell(14, 1, "Averages:").font = BOLD
    avg_rows = [
        (15, "Simple 3-yr",    "simple", 3),
        (16, "Wtd 3-yr",       "wtd",    3),
        (17, "Simple 5-yr",    "simple", 5),
        (18, "Wtd 5-yr",       "wtd",    5),
        (19, "Simple 7-yr",    "simple", 7),
        (20, "Wtd 7-yr",       "wtd",    7),
        (21, "Simple 10-yr",   "simple", 10),
        (22, "Wtd 10-yr",      "wtd",    10),
        (23, "Simple All-yrs", "simple", None),
        (24, "Wtd All-yrs",    "wtd",    None),
    ]
    for r, label, mode, n in avg_rows:
        ws.cell(r, 1, label).font = BOLD
        if not has_tri:
            continue
        for k in range(n_pairs):
            ldf_col = get_column_letter(2 + k)
            last_ratio_row = n_ays + 2 - 1 - k
            tri_last = n_ays + 1 - 1 - k
            if mode == "simple":
                start = 3 if n is None else max(3, last_ratio_row - n + 1)
                ws.cell(r, 2 + k,
                    f"=IFERROR(AVERAGE({ldf_col}{start}:{ldf_col}{last_ratio_row}),\"\")")
            else:
                start = 2 if n is None else max(2, tri_last - n + 1)
                f_col = get_column_letter(2 + k)
                t_col = get_column_letter(3 + k)
                ws.cell(r, 2 + k,
                    f"=IFERROR(SUM({tri_sheet}!{t_col}{start}:{t_col}{tri_last})"
                    f"/SUM({tri_sheet}!{f_col}{start}:{f_col}{tri_last}),\"\")")

    # Selected Age-to-Age (row 27); Tail from Inputs_Assumptions
    ws.cell(26, 1, "Selected Age-to-Age:").font = BOLD
    ws.cell(27, 1, "Selected").font = BOLD
    for k in range(n_pairs):
        col_l = get_column_letter(2 + k)
        ws.cell(27, 2 + k, f"={col_l}18" if has_tri else 1)
    ws.cell(27, tail_col, f"=Inputs_Assumptions!$B${src['TailRow']}")

    # Per-AY LDF table: ATAs first, cumulate only after selection
    ws.cell(29, 1, "Per-AY LDFs:").font = BOLD
    headers = ["AccidentYear", "CurrentDevAge", "Triangle ATA", "Industry ATA",
               "Broker ATA", "User ATA", "Selected ATA", "Cumulative LDF"]
    for j, h in enumerate(headers):
        c = ws.cell(30, j + 1, h); c.font = BOLD; c.fill = HDR_FILL

    # Assumes triangle eval months and AY position line up: oldest AY is at the
    # terminal age (uses the tail factor); each newer AY moves one step earlier.
    for i in range(n_ays):
        r = 31 + i
        in_row = 2 + i
        ws.cell(r, 1, f"=Inputs_Latest!A{in_row}").number_format = "0"
        # CurrentDevAge derived from position using the DevAges list
        ws.cell(r, 2, dev_ages[n_ays - 1 - i]).number_format = "0"
        # Triangle ATA: direct cell reference to Section 1 Selected by position
        section1_col = get_column_letter(2 + (n_ays - 1 - i))
        if has_tri:
            ws.cell(r, 3, f"={section1_col}27")
        # Industry / Broker / User ATAs straight from ldf_inputs
        ws.cell(r, 4, f"=Inputs_LDFs!{src['Industry']}{in_row}")
        ws.cell(r, 5, f"=Inputs_LDFs!{src['Broker']}{in_row}")
        ws.cell(r, 6, f"=Inputs_LDFs!{src['User']}{in_row}")
        # Selected ATA defaults to Triangle (or User if no triangle)
        ws.cell(r, 7, f"=C{r}" if has_tri else f"=F{r}")
        # Cumulative LDF = product of Selected ATAs from this row up through the
        # oldest AY (which holds the tail factor)
        ws.cell(r, 8, f"=PRODUCT(G$31:G{r})")
        for c in range(3, 9):
            ws.cell(r, c).number_format = "0.0000"

    # Widths
    ws.column_dimensions["A"].width = 24
    for c in range(2, tail_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    for r in range(3, 28):
        for c in range(2, tail_col + 1):
            ws.cell(r, c).number_format = "0.0000"


# ---------------------------------------------------------------------------
# LossRatios sheet
# ---------------------------------------------------------------------------
# Column layout (28 cols): A=AY, B=EP, C=OLF, D=OL_EP, E=PaidLoss, F=CumLDF_Paid,
# G=ReportedLoss, H=CumLDF_Reported, I=PaidLR, J=ReportedLR, K=Hist_CL_Paid,
# L=Hist_CL_Rep, M=LossTrend, N=ExpTrend, O=TrendYears, P=LossTF, Q=ExpTF,
# R=TrendedPaid, S=TrendedRep, T=TrendedOLEP, U=Prosp_CL_P, V=Prosp_CL_R,
# W=BF_P, X=BF_R, Y=UsedUp_P, Z=UsedUp_R, AA=CC_P, AB=CC_R
_LR_COLUMNS = [
    ("AccidentYear",               "=Inputs_Latest!A{r}",                                "0"),
    ("EarnedPremium",              "=Inputs_Latest!B{r}",                                "#,##0"),
    ("OnLevelFactor",              "=Inputs_Latest!C{r}",                                "0.0000"),
    ("OL_EP",                      "=B{r}*C{r}",                                         "#,##0"),
    ("PaidLoss",                   "=Inputs_Latest!D{r}",                                "#,##0"),
    ("CumLDF_Paid",                "=VLOOKUP(A{r},LDFs_Paid!LDF_RANGE,8,FALSE)",         "0.0000"),
    ("ReportedLoss",               "=Inputs_Latest!E{r}",                                "#,##0"),
    ("CumLDF_Reported",            "=VLOOKUP(A{r},LDFs_Reported!LDF_RANGE,8,FALSE)",     "0.0000"),
    ("PaidLR",                     "=E{r}/B{r}",                                         "0.00%"),
    ("ReportedLR",                 "=G{r}/B{r}",                                         "0.00%"),
    ("Hist_CL_Paid_ULR",           "=E{r}*F{r}/B{r}",                                    "0.00%"),
    ("Hist_CL_Reported_ULR",       "=G{r}*H{r}/B{r}",                                    "0.00%"),
    ("LossTrend",                  "=Inputs_Latest!F{r}",                                "0.00%"),
    ("ExposureTrend",              "=Inputs_Latest!G{r}",                                "0.00%"),
    ("TrendYears",                 "=(Inputs_Assumptions!$B$2-DATE(A{r},7,1))/365.25",   "0.00"),
    ("LossTrendFactor",            "=(1+M{r})^O{r}",                                     "0.0000"),
    ("ExpTrendFactor",             "=(1+N{r})^O{r}",                                     "0.0000"),
    ("TrendedPaid",                "=E{r}*P{r}",                                         "#,##0"),
    ("TrendedReported",            "=G{r}*P{r}",                                         "#,##0"),
    ("TrendedOL_EP",               "=D{r}*Q{r}",                                         "#,##0"),
    ("Prosp_CL_Paid_ULR_TOL",      "=R{r}*F{r}/T{r}",                                    "0.00%"),
    ("Prosp_CL_Reported_ULR_TOL",  "=S{r}*H{r}/T{r}",                                    "0.00%"),
    ("BF_Paid_ULR_TOL",            "=R{r}/T{r}+$U$AXMRY*(1-1/F{r})",                     "0.00%"),
    ("BF_Reported_ULR_TOL",        "=S{r}/T{r}+$V$AXMRY*(1-1/H{r})",                     "0.00%"),
    ("UsedUp_Paid",                "=T{r}/F{r}",                                         "#,##0"),
    ("UsedUp_Reported",            "=T{r}/H{r}",                                         "#,##0"),
    ("CC_Paid_ULR_TOL",            "=R{r}/T{r}+($R$AXMRY/$Y$AXMRY)*(1-1/F{r})",          "0.00%"),
    ("CC_Reported_ULR_TOL",        "=S{r}/T{r}+($S$AXMRY/$Z$AXMRY)*(1-1/H{r})",          "0.00%"),
]

# Totals row specs: keyed by column letter; value tells how to aggregate.
#   "sum"                                  -> SUM
#   ("ratio_sum", num_col, den_col)        -> SUM(num)/SUM(den)
#   ("ratio_sp",  n1, n2, den)             -> SUMPRODUCT(n1,n2)/SUM(den)
_TOTALS = {
    "B":  "sum",
    "D":  "sum",
    "E":  "sum",
    "G":  "sum",
    "I":  ("ratio_sum", "E", "B"),
    "J":  ("ratio_sum", "G", "B"),
    "K":  ("ratio_sp",  "E", "F", "B"),
    "L":  ("ratio_sp",  "G", "H", "B"),
    "R":  "sum",
    "S":  "sum",
    "T":  "sum",
    "U":  ("ratio_sp",  "R", "F", "T"),
    "V":  ("ratio_sp",  "S", "H", "T"),
    "W":  ("ratio_sp",  "W", "T", "T"),
    "X":  ("ratio_sp",  "X", "T", "T"),
    "Y":  "sum",
    "Z":  "sum",
    "AA": ("ratio_sp",  "AA", "T", "T"),
    "AB": ("ratio_sp",  "AB", "T", "T"),
}


def _totals_formula(spec, col, start, end):
    if spec == "sum":
        return f"=SUM({col}{start}:{col}{end})"
    typ = spec[0]
    if typ == "ratio_sum":
        _, n, d = spec
        return f"=SUM({n}{start}:{n}{end})/SUM({d}{start}:{d}{end})"
    _, n1, n2, d = spec  # ratio_sp
    return f"=SUMPRODUCT({n1}{start}:{n1}{end},{n2}{start}:{n2}{end})/SUM({d}{start}:{d}{end})"


def _build_loss_ratios_sheet(wb, latest):
    ws = wb.create_sheet("LossRatios")
    n = len(latest)
    last = 1 + n
    ldf_range = f"$A$31:$H${30 + n}"
    totals_label_row = last + 3
    rows_3 = totals_label_row + 1
    rows_5 = totals_label_row + 2
    rows_7 = totals_label_row + 3
    rows_10 = totals_label_row + 4
    rows_all = totals_label_row + 5
    rows_axmry = totals_label_row + 6
    # BF/CC anchor falls back to All Years if n < 3 (need >= 2 AYs ex-MRY)
    anchor_row = rows_axmry if n >= 3 else rows_all

    def _subst(f):
        return f.replace("LDF_RANGE", ldf_range).replace("AXMRY", str(anchor_row))

    # Headers
    for j, (h, _, _) in enumerate(_LR_COLUMNS):
        c = ws.cell(1, j + 1, h); c.font = BOLD; c.fill = HDR_FILL

    # Per-AY data
    for i in range(n):
        r = 2 + i
        for j, (_, tmpl, fmt) in enumerate(_LR_COLUMNS):
            c = ws.cell(r, j + 1, _subst(tmpl).format(r=r))
            c.number_format = fmt

    # Totals section — each window has a min_n; rows are blank if n is too small
    ws.cell(totals_label_row, 1, "Totals:").font = BOLD
    windows = [
        (rows_3,    "3-yr",          3,  max(2, last - 2),  last),
        (rows_5,    "5-yr",          5,  max(2, last - 4),  last),
        (rows_7,    "7-yr",          7,  max(2, last - 6),  last),
        (rows_10,   "10-yr",         10, max(2, last - 9),  last),
        (rows_all,  "All Years",     1,  2,                 last),
        (rows_axmry,"All Yrs x MRY", 3,  2,                 last - 1),
    ]
    for r, label, min_n, start, end in windows:
        ws.cell(r, 1, label).font = BOLD
        if n < min_n:
            continue
        for j, (_, _, fmt) in enumerate(_LR_COLUMNS):
            col_letter = get_column_letter(j + 1)
            spec = _TOTALS.get(col_letter)
            if spec is None:
                continue
            cell = ws.cell(r, j + 1, _totals_formula(spec, col_letter, start, end))
            cell.number_format = fmt
            cell.font = BOLD

    # Widths
    ws.column_dimensions["A"].width = 12
    for j in range(1, len(_LR_COLUMNS)):
        ws.column_dimensions[get_column_letter(j + 1)].width = 13
    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------
def _build_summary_sheet(wb, latest):
    ws = wb.create_sheet("Summary")
    headers = [
        "AccidentYear",
        "Hist CL Paid", "Hist CL Reported",
        "Prosp CL Paid TOL", "Prosp CL Reported TOL",
        "Prosp BF Paid TOL", "Prosp BF Reported TOL",
        "Prosp CC Paid TOL", "Prosp CC Reported TOL",
    ]
    for j, h in enumerate(headers):
        c = ws.cell(1, j + 1, h); c.font = BOLD; c.fill = HDR_FILL

    # LossRatios source columns
    src = ["K", "L", "U", "V", "W", "X", "AA", "AB"]
    n = len(latest)
    last = 1 + n
    totals_label_row = last + 3

    for i in range(n):
        r = 2 + i
        ws.cell(r, 1, f"=LossRatios!A{r}").number_format = "0"
        for j, lr_col in enumerate(src):
            c = ws.cell(r, 2 + j, f"=LossRatios!{lr_col}{r}")
            c.number_format = "0.00%"

    # Totals — reference LossRatios totals row-by-row; skip if n too small
    ws.cell(totals_label_row, 1, "Totals:").font = BOLD
    labels = [("3-yr", 3), ("5-yr", 5), ("7-yr", 7), ("10-yr", 10),
              ("All Years", 1), ("All Yrs x MRY", 3)]
    for w_idx, (label, min_n) in enumerate(labels):
        r = totals_label_row + 1 + w_idx
        ws.cell(r, 1, label).font = BOLD
        if n < min_n:
            continue
        for j, lr_col in enumerate(src):
            c = ws.cell(r, 2 + j, f"=LossRatios!{lr_col}{r}")
            c.number_format = "0.00%"; c.font = BOLD

    ws.column_dimensions["A"].width = 13
    for j in range(2, 10):
        ws.column_dimensions[get_column_letter(j)].width = 19


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------
def build_analysis(input_path=None, output_path=None):
    input_path = Path(input_path or _path())
    output_path = Path(output_path or input_path.parent / ANALYSIS_WORKBOOK)
    latest, triangles, ldf_inputs, asmp = load_all(input_path)

    dev_ages = asmp.get("DevAges") or [12, 24, 36, 48, 60, 72, 84, 96, 108, 120]
    n_ays, n_dev = len(latest), len(dev_ages)
    warnings: list[str] = []
    if n_ays != n_dev:
        n = min(n_ays, n_dev)
        if n_ays > n:
            warnings.append(
                f"latest_data has {n_ays} AYs but DevAges has {n_dev} entries; "
                f"dropping the {n_ays - n} oldest AY(s) and using {n} of each.")
            latest = latest.tail(n).reset_index(drop=True)
            ldf_inputs = ldf_inputs.tail(n).reset_index(drop=True)
        else:
            warnings.append(
                f"DevAges has {n_dev} entries but latest_data has only {n_ays} AYs; "
                f"dropping the {n_dev - n} largest dev age(s) and using {n} of each.")
            dev_ages = dev_ages[:n]

    if len(latest) < 3:
        warnings.append(
            f"Only {len(latest)} AY(s) supplied; BF a priori and CC ELR fall back "
            f"to the All Years total (All Yrs x MRY needs at least 3 AYs).")

    for w in warnings:
        print("Warning:", w)

    wb = Workbook()
    wb.remove(wb.active)
    _write_inputs(wb, latest, ldf_inputs, asmp, triangles)
    if warnings:
        ws = wb["Inputs_Assumptions"]
        for w in warnings:
            r = ws.max_row + 2
            ws.cell(r, 1, "WARNING").font = BOLD
            c = ws.cell(r, 2, w)
            c.fill = PatternFill("solid", fgColor="FFFF00")
    has_paid     = not triangles.empty and (triangles["type"] == "paid").any()
    has_reported = not triangles.empty and (triangles["type"] == "reported").any()
    _build_ldf_sheet(wb, "paid", latest, has_paid, dev_ages)
    _build_ldf_sheet(wb, "reported", latest, has_reported, dev_ages)
    _build_loss_ratios_sheet(wb, latest)
    _build_summary_sheet(wb, latest)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# PDF report
# ---------------------------------------------------------------------------
def build_pdf_report(project_dir, analysis_path=None):
    """Render a one-page PDF summary from the analysis workbook.

    Returns the PDF path on success, or ``None`` if LibreOffice is unavailable
    (the PDF leg needs it to evaluate the workbook's formulas).
    """
    import shutil
    import subprocess
    import tempfile
    from datetime import datetime
    from openpyxl import load_workbook
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    project_dir = Path(project_dir)
    analysis_path = Path(analysis_path or project_dir / ANALYSIS_WORKBOOK)
    pdf_path = project_dir / REPORT_PDF

    soffice = shutil.which("libreoffice") or shutil.which("soffice")
    if not soffice:
        print("Warning: libreoffice/soffice not found on PATH; skipping PDF report.")
        return None

    # Round-trip through LibreOffice to evaluate the workbook's formulas.
    with tempfile.TemporaryDirectory() as td:
        result = subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             str(analysis_path), "--outdir", td],
            capture_output=True)
        evaluated = Path(td) / analysis_path.name
        if result.returncode != 0 or not evaluated.exists():
            print(f"Warning: LibreOffice failed to evaluate workbook; skipping PDF report.\n"
                  f"  stderr: {result.stderr.decode(errors='replace')[:200]}")
            return None
        wb = load_workbook(evaluated, data_only=True)

    ws = wb["Summary"]
    headers = [ws.cell(1, c).value for c in range(1, 10)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 10)]
        if any(v is not None for v in row):
            rows.append(row)

    ws_a = wb["Inputs_Assumptions"]
    warnings = [ws_a.cell(r, 2).value for r in range(1, ws_a.max_row + 1)
                if ws_a.cell(r, 1).value == "WARNING"]
    notes = _build_notes(rows, warnings)

    def fmt(v):
        if v is None:
            return ""
        if isinstance(v, float):
            return f"{v:.1%}"
        return str(v)

    data = [headers] + [[fmt(v) for v in row] for row in rows]
    doc = SimpleDocTemplate(
        str(pdf_path), pagesize=landscape(letter),
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    flow = [
        Paragraph("Actuarial Pricing Analysis", styles["Title"]),
        Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d')}", styles["Normal"]),
        Spacer(1, 0.2 * inch),
        Paragraph("Summary", styles["Heading2"]),
    ]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
    ]))
    flow.append(table)
    flow.append(Spacer(1, 0.2 * inch))
    flow.append(Paragraph("Observations", styles["Heading2"]))
    for note in notes:
        flow.append(Paragraph(f"&bull; {note}", styles["Normal"]))
    doc.build(flow)
    return pdf_path


def _build_notes(rows, warnings):
    """Return a short list of observations and warnings to include in the PDF."""
    notes = list(warnings)
    ay_rows = [r for r in rows if isinstance(r[0], int)]
    if not ay_rows:
        return notes or ["No accident-year data."]

    cl_p = [r[3] for r in ay_rows if isinstance(r[3], (int, float))]
    if len(cl_p) >= 3:
        spread = max(cl_p) - min(cl_p)
        if spread > 0.10:
            notes.append(
                f"Prospective CL paid ULR varies by {spread:.1%} across AYs — "
                "year-over-year variation is high.")

    latest = ay_rows[-1]
    if isinstance(latest[3], (int, float)) and isinstance(latest[5], (int, float)):
        diff = latest[5] - latest[3]
        if abs(diff) > 0.05:
            notes.append(
                f"For AY {latest[0]}, BF paid ULR is {diff:+.1%} versus CL — "
                "BF is reweighting heavily toward the a priori.")

    if any(isinstance(v, (int, float)) and v < 0 for r in ay_rows for v in r[1:]):
        notes.append("Negative ULR detected — check loss data for sign errors.")

    if not notes:
        notes.append("No unusual patterns detected.")
    return notes


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Build actuarial pricing analysis from pricing_inputs.xlsx.")
    parser.add_argument(
        "project_dir", nargs="?", type=Path,
        help="Directory containing pricing_inputs.xlsx (defaults to script directory).")
    args = parser.parse_args()
    project_dir = args.project_dir or Path(__file__).resolve().parent

    out_xlsx = build_analysis(
        input_path=project_dir / WORKBOOK,
        output_path=project_dir / ANALYSIS_WORKBOOK,
    )
    print(f"wrote {out_xlsx}")
    out_pdf = build_pdf_report(project_dir, analysis_path=out_xlsx)
    if out_pdf:
        print(f"wrote {out_pdf}")
